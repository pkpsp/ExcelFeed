import openpyxl 

class UpdateExcel():

    excelFile = "demo excel.xlsx"

    def __init__(self):
        # super().__init__()
        self.openExcel()

    def openExcel(self):
        excelObj = openpyxl.load_workbook(excelFile) 
        self.excelSheetObj = excelObj.active 

    def closeExcel(self):
        if self.excelSheetObj != None:
            self.excelSheetObj.save(excelFile)
            self.excelSheetObj = None

    def getCell(self, cellOrRow, column = -1):
        if column > 0:
            return self.excelSheetObj.cell(row = cellOrRow, column = column).value
        else:
            return self.excelSheetObj[cellOrRow].value

    def putCell(self, value, cellOrRow, column = -1):
        try:
            if column > 0:
                self.excelSheetObj.cell(row = cellOrRow, column = column).value = value
                return True
            else:
                self.excelSheetObj[cellOrRow] = value
                return True
        except:
            return False